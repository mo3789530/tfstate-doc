# create write csv or excel format
import openpyxl

class ExcelWriter:
    workbook = None
    def __init__(self) -> None:
        self.workbook = openpyxl.Workbook()

    def save_workbook(self, name):
        path = "" + "./" + name + ".xlsx"
        self.workbook.save(path)

    def create_sheet(self, name):
        if self.is_created_sheet(name) == False:
            self.workbook.create_sheet(title=name)

    def is_created_sheet(self, name) -> bool:
        try:
            self.workbook[name]
            return True
        except:
            return False

    # def write_sheet(self, name, dic):
    #     if self.is_created_sheet(name) == False:
    #         self.workbook.create_sheet(title=name)
    #     sheet = self.workbook[name]
    #     for key, value in dic.items():
    #         sheet[key] = value

    def write_sheet(self, name, dic):
        index = 0
        if self.is_created_sheet(name) == False:
            self.workbook.create_sheet(title=name)
        sheet = self.workbook[name]
        index = sheet.max_row + 1
        for key, value in dic.items():
            sheet.cell(row=index, column=1).value = key
            sheet.cell(row=index, column=2).value = value